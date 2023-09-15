import instaloader
from openpyxl import Workbook

# Crie uma instância do Instaloader
loader = instaloader.Instaloader()

# Nome do usuário
nome_do_usuario = 'seplag.pe'

# Carregue um perfil público
perfilFafire = instaloader.Profile.from_username(loader.context, nome_do_usuario)

# Crie uma nova planilha
workbook = Workbook()
sheet = workbook.active

# Crie cabeçalhos para as colunas
sheet['A1'] = 'Título'
sheet['B1'] = 'Data do Post'
sheet['C1'] = 'Hora do Post'
sheet['D1'] = 'Patrocínio'
sheet['E1'] = 'Curtidas'
sheet['F1'] = 'Comentários'
sheet['G1'] = 'Visualizações'
sheet['H1'] = 'URL'

# Itere pelas postagens e adicione os dados à planilha
row = 2  # Iniciar na segunda linha após o cabeçalho
for post in perfilFafire.get_posts():
    sheet.cell(row=row, column=1, value=post.caption)
    sheet.cell(row=row, column=2, value=post.date.strftime('%Y-%m-%d'))
    sheet.cell(row=row, column=3, value=post.date.strftime('%H:%M:%S'))
    sheet.cell(row=row, column=4, value=post.is_sponsored)
    sheet.cell(row=row, column=5, value=post.likes)
    sheet.cell(row=row, column=6, value=post.comments)
    sheet.cell(row=row, column=7, value=post.video_view_count)
    sheet.cell(row=row, column=8, value=post.url)    
    
    row += 1

# Salve a planilha em um arquivo Excel
workbook.save(f'instagram_{nome_do_usuario}.xlsx')